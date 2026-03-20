import os
import sqlite3
import traceback
from datetime import UTC, datetime, timedelta

import discord
from discord import ui
from discord.ext import commands
from openpyxl import Workbook, load_workbook


# ========================= CONFIG =========================
TOKEN = os.getenv("DISCORD_BOT_TOKEN", "MTQ3MDc4MjMwOTgzNzA0NTk1Nw.GABdTC.jhqnUqfRStNMowLEX_1l3oUqEpI92AL6udI3hs")
ROLE_LEADER_ID = 1390977162164568105, 1390977162198384685, 1390977162198384686, 1390977162173091949
CHANNEL_REPORT_ID = 1390994169530552422
CHANNEL_VERIFY_ID = 1469371502766981414
DELETE_AFTER = 10

DB_FILE = "database.db"
EXCEL_FILE = "logs.xlsx"


# ========================= CONTRACT SYSTEM =========================
CONTRACT_DATA = {
    "Гровер 3": {1: 140000, 2: 70000, 3: 50000, 4: 30000},
    "Гровер 2": {1: 120000, 2: 60000, 3: 40000, 4: 30000},
    "Гровер 1": {1: 60000, 2: 30000},
    "Майстер баків 2": {2: 60000},
    "Майстер баків 1": {2: 50000},
    "Доставка OG Kush": {1: 80000, 2: 40000, 3: 30000, 4: 20000},
    "На буксир 2": {2: 60000, 4: 30000},
    "На буксир 1": {2: 40000, 4: 20000},
    "Балоний транзит 1": {1: 80000, 2: 40000, 3: 20000, 4: 20000},
    "Балоний транзит 2": {1: 100000, 2: 50000, 3: 30000, 4: 20000},
    "Під корень 1": {1: 70000, 2: 30000, 3: 20000, 4: 20000},
    "Під корень 2": {1: 90000, 2: 50000, 3: 30000, 4: 20000},
}


# ========================= EXCEL =========================
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Дата", "Нік", "Контракти", "Заробіток", "Штраф", "Виплата"])
    wb.save(EXCEL_FILE)


def log_to_excel(data):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in data:
        ws.append(row)
    wb.save(EXCEL_FILE)


# ========================= DATABASE =========================
conn = sqlite3.connect(DB_FILE, check_same_thread=False)
c = conn.cursor()

TABLE_SCHEMAS = {
    "pending_reports": """
        CREATE TABLE pending_reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            author TEXT,
            contract TEXT,
            participants TEXT
        )
    """,
    "contracts": """
        CREATE TABLE contracts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            participant TEXT,
            amount INTEGER,
            timestamp TEXT
        )
    """,
    "fines": """
        CREATE TABLE fines (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user TEXT,
            amount INTEGER,
            reason TEXT,
            timestamp TEXT
        )
    """,
}

REQUIRED_COLUMNS = {
    "pending_reports": {"id", "author", "contract", "participants"},
    "contracts": {"id", "participant", "amount", "timestamp"},
    "fines": {"id", "user", "amount", "reason", "timestamp"},
}


def ensure_table_schema(table_name: str):
    c.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name = ?",
        (table_name,),
    )
    exists = c.fetchone() is not None

    if not exists:
        c.execute(TABLE_SCHEMAS[table_name])
        return

    c.execute(f"PRAGMA table_info({table_name})")
    current_columns = {row[1] for row in c.fetchall()}

    if REQUIRED_COLUMNS[table_name].issubset(current_columns):
        return

    backup_name = f"{table_name}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    c.execute(f"ALTER TABLE {table_name} RENAME TO {backup_name}")
    c.execute(TABLE_SCHEMAS[table_name])
    print(
        f"Database schema mismatch for '{table_name}'. "
        f"Created fresh table and kept backup as '{backup_name}'."
    )


for table in ("pending_reports", "contracts", "fines"):
    ensure_table_schema(table)

conn.commit()


# ========================= BOT =========================
intents = discord.Intents.default()
bot = commands.Bot(command_prefix="!", intents=intents)


def has_leader_role(interaction: discord.Interaction) -> bool:
    return isinstance(interaction.user, discord.Member) and any(
        role.id == ROLE_LEADER_ID for role in interaction.user.roles
    )


class SafeView(ui.View):
    async def on_error(self, interaction: discord.Interaction, error: Exception, item):
        print("VIEW ERROR:", repr(error))
        traceback.print_exc()

        if interaction.response.is_done():
            await interaction.followup.send(
                "❌ Сталася помилка. Перевір консоль бота.",
                ephemeral=True,
                delete_after=DELETE_AFTER,
            )
        else:
            await interaction.response.send_message(
                "❌ Сталася помилка. Перевір консоль бота.",
                ephemeral=True,
                delete_after=DELETE_AFTER,
            )


# ========================= REPORT FLOW =========================
class ContractSelect(ui.Select):
    def __init__(self, author: str):
        options = [discord.SelectOption(label=name) for name in CONTRACT_DATA.keys()]
        super().__init__(placeholder="Оберіть контракт", options=options)
        self.author = author

    async def callback(self, interaction: discord.Interaction):
        contract = self.values[0]
        view = PeopleView(self.author, contract)

        await interaction.response.edit_message(
            content=f"Контракт: **{contract}**\nОберіть кількість людей",
            view=view,
        )


class ContractView(SafeView):
    def __init__(self, author: str):
        super().__init__(timeout=300)
        self.add_item(ContractSelect(author))


class PeopleSelect(ui.Select):
    def __init__(self, author: str, contract: str):
        options = [discord.SelectOption(label=str(i)) for i in range(1, 7)]
        super().__init__(
            placeholder="Скільки людей виконувало контракт",
            options=options,
        )
        self.author = author
        self.contract = contract

    async def callback(self, interaction: discord.Interaction):
        people = int(self.values[0])

        if people not in CONTRACT_DATA[self.contract]:
            return await interaction.response.send_message(
                "❌ Для цього контракту така кількість людей недоступна",
                ephemeral=True,
                delete_after=DELETE_AFTER,
            )

        view = ConfirmView(self.author, self.contract, people)
        await interaction.response.edit_message(
            content=f"Контракт: {self.contract}\nЛюдей: {people}",
            view=view,
        )


class PeopleView(SafeView):
    def __init__(self, author: str, contract: str):
        super().__init__(timeout=300)
        self.add_item(PeopleSelect(author, contract))


class ConfirmView(SafeView):
    def __init__(self, author: str, contract: str, people: int):
        super().__init__(timeout=300)
        self.author = author
        self.contract = contract
        self.people = people

    @ui.button(label="➡ Другий етап", style=discord.ButtonStyle.success)
    async def next_step(self, interaction: discord.Interaction, button: ui.Button):
        await interaction.response.send_modal(
            FinalModal(self.author, self.contract, self.people)
        )


class FinalModal(ui.Modal, title="Фінальний етап"):
    def __init__(self, author: str, contract: str, people: int):
        super().__init__()
        self.author = author
        self.contract = contract
        self.people = people

        self.participants = ui.TextInput(
            label="Учасники через кому",
            placeholder="Ivan, Petro, Andriy",
        )
        self.screenshot_1 = ui.TextInput(
            label="Скріншот 1 (посилання)",
            placeholder="https://...",
        )
        self.screenshot_2 = ui.TextInput(
            label="Скріншот 2 (посилання)",
            placeholder="https://...",
        )
        self.add_item(self.participants)
        self.add_item(self.screenshot_1)
        self.add_item(self.screenshot_2)

    async def on_submit(self, interaction: discord.Interaction):
        names = [x.strip() for x in self.participants.value.split(",") if x.strip()]

        if len(names) != self.people:
            return await interaction.response.send_message(
                f"❌ Має бути {self.people} людей",
                ephemeral=True,
                delete_after=DELETE_AFTER,
            )

        if self.people == 1 and "," in self.participants.value:
            return await interaction.response.send_message(
                "❌ Для 1 людини не потрібно ком",
                ephemeral=True,
                delete_after=DELETE_AFTER,
            )

        if self.people > 1 and self.participants.value.count(",") != self.people - 1:
            return await interaction.response.send_message(
                f"❌ Має бути {self.people - 1} ком",
                ephemeral=True,
                delete_after=DELETE_AFTER,
            )

        if len(names) != len(set(names)):
            return await interaction.response.send_message(
                "❌ Є дублікати учасників",
                ephemeral=True,
                delete_after=DELETE_AFTER,
            )

        screenshot_1 = self.screenshot_1.value.strip()
        screenshot_2 = self.screenshot_2.value.strip()
        screenshot_links = [screenshot_1, screenshot_2]

        for link in screenshot_links:
            if not (link.startswith("http://") or link.startswith("https://")):
                return await interaction.response.send_message(
                    "❌ Скріншоти мають бути у вигляді посилань http/https",
                    ephemeral=True,
                    delete_after=DELETE_AFTER,
                )

        if self.people not in CONTRACT_DATA[self.contract]:
            return await interaction.response.send_message(
                "❌ Немає виплати для цієї кількості людей",
                ephemeral=True,
                delete_after=DELETE_AFTER,
            )

        price = CONTRACT_DATA[self.contract][self.people]
        payout_text = "\n".join(f"{name} — {price}$" for name in names)

        c.execute(
            "INSERT INTO pending_reports(author, contract, participants) VALUES(?, ?, ?)",
            (self.author, self.contract, self.participants.value),
        )
        conn.commit()

        report_id = c.lastrowid
        channel = bot.get_channel(CHANNEL_VERIFY_ID)
        if channel is None:
            return await interaction.response.send_message(
                "❌ Канал для перевірки не знайдено",
                ephemeral=True,
                delete_after=DELETE_AFTER,
            )

        embed = discord.Embed(title="📝 Новий рапорт на премію", color=0x2ECC71)
        embed.add_field(name="Хто подав", value=self.author, inline=False)
        embed.add_field(name="Контракт", value=self.contract, inline=False)
        embed.add_field(name="Людей", value=str(self.people), inline=False)
        embed.add_field(name="Учасники", value="\n".join(names), inline=False)
        embed.add_field(name="Виплата", value=payout_text, inline=False)
        embed.add_field(
            name="Скріншоти",
            value=f"[Скріншот 1]({screenshot_1})\n[Скріншот 2]({screenshot_2})",
            inline=False,
        )
        embed.add_field(name="Статус", value="⏳ Очікує перевірки", inline=False)

        screenshot_embed_1 = discord.Embed(color=0x2ECC71)
        screenshot_embed_1.set_image(url=screenshot_1)

        screenshot_embed_2 = discord.Embed(color=0x2ECC71)
        screenshot_embed_2.set_image(url=screenshot_2)

        await channel.send(
            embeds=[embed, screenshot_embed_1, screenshot_embed_2],
            view=VerifyView(report_id),
        )
        await interaction.response.send_message(
            "✅ Рапорт відправлено",
            ephemeral=True,
            delete_after=DELETE_AFTER,
        )


# ========================= FINE FLOW =========================
class FineModal(ui.Modal, title="Видати штраф"):
    user = ui.TextInput(label="Нік + статік", placeholder="Romeo Hunter #19857")
    amount = ui.TextInput(label="Сума штрафу", placeholder="5000")
    reason = ui.TextInput(label="Причина")

    async def on_submit(self, interaction: discord.Interaction):
        if not self.amount.value.isdigit():
            return await interaction.response.send_message(
                "❌ Сума має бути числом",
                ephemeral=True,
                delete_after=DELETE_AFTER,
            )

        channel = bot.get_channel(CHANNEL_VERIFY_ID)
        if channel is None:
            return await interaction.response.send_message(
                "❌ Канал не знайдено",
                ephemeral=True,
                delete_after=DELETE_AFTER,
            )

        embed = discord.Embed(title="⚠ Новий штраф", color=0xE74C3C)
        embed.add_field(name="Порушник", value=self.user.value, inline=False)
        embed.add_field(name="Сума", value=f"{self.amount.value}$", inline=False)
        embed.add_field(name="Причина", value=self.reason.value, inline=False)
        embed.add_field(
            name="Статус",
            value="⏳ Очікує на перевірку лідера",
            inline=False,
        )

        await channel.send(
            embed=embed,
            view=FineVerifyView(
                self.user.value, int(self.amount.value), self.reason.value
            ),
        )
        await interaction.response.send_message(
            "✅ Відправлено на перевірку лідеру",
            ephemeral=True,
            delete_after=DELETE_AFTER,
        )


# ========================= MAIN VIEW =========================
class MainView(SafeView):
    def __init__(self):
        super().__init__(timeout=None)

    @ui.button(label="📨 Подати звіт", style=discord.ButtonStyle.primary)
    async def submit(self, interaction: discord.Interaction, button: ui.Button):
        author_name = (
            interaction.user.display_name
            if isinstance(interaction.user, discord.Member)
            else interaction.user.name
        )
        await interaction.response.send_message(
            "👇 Оберіть контракт",
            view=ContractView(author_name),
            ephemeral=True,
        )

    @ui.button(label="⚠ Видати штраф", style=discord.ButtonStyle.danger)
    async def fine(self, interaction: discord.Interaction, button: ui.Button):
        if not has_leader_role(interaction):
            return await interaction.response.send_message(
                "❌ У тебе немає доступу до видачі штрафів",
                ephemeral=True,
                delete_after=DELETE_AFTER,
            )
        await interaction.response.send_modal(FineModal())


# ========================= VERIFY REPORT =========================
class VerifyView(SafeView):
    STATUS_FIELD_INDEX = 6

    def __init__(self, report_id: int):
        super().__init__(timeout=None)
        self.report_id = report_id

    async def finish(self, interaction: discord.Interaction, status: str):
        embed = interaction.message.embeds[0]
        embed.set_field_at(
            self.STATUS_FIELD_INDEX,
            name="Статус",
            value=f"{status} {interaction.user.mention}",
            inline=False,
        )

        for item in self.children:
            item.disabled = True

        await interaction.message.edit(embed=embed, view=self)

    @ui.button(label="✅ Схвалити", style=discord.ButtonStyle.success)
    async def approve(self, interaction: discord.Interaction, button: ui.Button):
        if not has_leader_role(interaction):
            return await interaction.response.send_message(
                "❌ Немає доступу",
                ephemeral=True,
                delete_after=DELETE_AFTER,
            )

        c.execute(
            "SELECT author, participants FROM pending_reports WHERE id = ?",
            (self.report_id,),
        )
        row = c.fetchone()
        if not row:
            return await interaction.response.send_message(
                "⚠ Уже оброблено",
                ephemeral=True,
                delete_after=DELETE_AFTER,
            )

        _, participants = row
        users = [x.strip() for x in participants.split(",") if x.strip()]

        contract_name = interaction.message.embeds[0].fields[1].value
        people_count = int(interaction.message.embeds[0].fields[2].value)
        price = CONTRACT_DATA[contract_name][people_count]

        for user in set(users):
            c.execute(
                "INSERT INTO contracts(participant, amount, timestamp) VALUES(?, ?, ?)",
                (user, price, datetime.now(UTC).isoformat()),
            )

        c.execute("DELETE FROM pending_reports WHERE id = ?", (self.report_id,))
        conn.commit()

        await self.finish(interaction, "✅ Схвалено")
        await interaction.response.send_message(
            "✅ Звіт схвалено",
            ephemeral=True,
            delete_after=DELETE_AFTER,
        )

    @ui.button(label="❌ Відмовити", style=discord.ButtonStyle.danger)
    async def reject(self, interaction: discord.Interaction, button: ui.Button):
        if not has_leader_role(interaction):
            return await interaction.response.send_message(
                "❌ Немає доступу",
                ephemeral=True,
                delete_after=DELETE_AFTER,
            )

        c.execute("DELETE FROM pending_reports WHERE id = ?", (self.report_id,))
        conn.commit()

        await self.finish(interaction, "❌ Відхилено")
        await interaction.response.send_message(
            "❌ Звіт відхилено",
            ephemeral=True,
            delete_after=DELETE_AFTER,
        )


# ========================= VERIFY FINE =========================
class FineVerifyView(SafeView):
    STATUS_FIELD_INDEX = 3

    def __init__(self, user: str, amount: int, reason: str):
        super().__init__(timeout=None)
        self.user = user
        self.amount = amount
        self.reason = reason

    async def finish(self, interaction: discord.Interaction, status: str):
        embed = interaction.message.embeds[0]
        embed.set_field_at(
            self.STATUS_FIELD_INDEX,
            name="Статус",
            value=f"{status} {interaction.user.mention}",
            inline=False,
        )

        for item in self.children:
            item.disabled = True

        await interaction.message.edit(embed=embed, view=self)

    @ui.button(label="✅ Схвалити", style=discord.ButtonStyle.success)
    async def approve(self, interaction: discord.Interaction, button: ui.Button):
        if not has_leader_role(interaction):
            return await interaction.response.send_message(
                "❌ Немає доступу",
                ephemeral=True,
                delete_after=DELETE_AFTER,
            )

        c.execute(
            "INSERT INTO fines(user, amount, reason, timestamp) VALUES(?, ?, ?, ?)",
            (self.user, self.amount, self.reason, datetime.now(UTC).isoformat()),
        )
        conn.commit()

        await self.finish(interaction, "✅ Схвалено")
        await interaction.response.send_message(
            "✅ Штраф схвалено",
            ephemeral=True,
            delete_after=DELETE_AFTER,
        )

    @ui.button(label="❌ Відмовити", style=discord.ButtonStyle.danger)
    async def reject(self, interaction: discord.Interaction, button: ui.Button):
        if not has_leader_role(interaction):
            return await interaction.response.send_message(
                "❌ Немає доступу",
                ephemeral=True,
                delete_after=DELETE_AFTER,
            )

        await self.finish(interaction, "❌ Відхилено")
        await interaction.response.send_message(
            "❌ Штраф відхилено",
            ephemeral=True,
            delete_after=DELETE_AFTER,
        )


# ========================= COMMANDS =========================
@bot.tree.command(name="weekly_report", description="Щотижневий фінансовий звіт")
async def weekly_report(interaction: discord.Interaction):
    week_ago = datetime.now() - timedelta(days=7)

    c.execute(
        """
        SELECT participant, COUNT(*), SUM(amount)
        FROM contracts
        WHERE timestamp >= ?
        GROUP BY participant
        """,
        (week_ago.isoformat(),),
    )
    contracts = c.fetchall()

    if not contracts:
        return await interaction.response.send_message(
            "❌ Немає даних",
            ephemeral=True,
            delete_after=DELETE_AFTER,
        )

    c.execute(
        "SELECT user, SUM(amount) FROM fines WHERE timestamp >= ? GROUP BY user",
        (week_ago.isoformat(),),
    )
    fines_data = dict(c.fetchall())

    embed = discord.Embed(title="📊 Щотижневий фінансовий звіт", color=0x2ECC71)
    excel_rows = []
    total = 0

    for user, count, salary in contracts:
        salary = salary or 0
        fine = fines_data.get(user, 0)
        payout = salary - fine
        total += payout

        embed.add_field(
            name=user,
            value=(
                f"Контракти: **{count}**\n"
                f"Заробіток: **${salary}**\n"
                f"Штраф: **${fine}**\n"
                f"Виплата: **${payout}**"
            ),
            inline=False,
        )

        excel_rows.append(
            [
                datetime.now().strftime("%Y-%m-%d %H:%M"),
                user,
                count,
                salary,
                fine,
                payout,
            ]
        )

    embed.set_footer(text=f"Загальна виплата: ${total}")

    await interaction.channel.send(embed=embed)
    await interaction.response.send_message(
        "✅ Звіт сформовано та збережено",
        ephemeral=True,
        delete_after=DELETE_AFTER,
    )

    log_to_excel(excel_rows)
    c.execute("DELETE FROM contracts")
    c.execute("DELETE FROM fines")
    conn.commit()


@bot.tree.command(name="setup", description="Створити панель")
async def setup(interaction: discord.Interaction):
    if interaction.channel_id != CHANNEL_REPORT_ID:
        return await interaction.response.send_message(
            "❌ Не той канал",
            ephemeral=True,
            delete_after=DELETE_AFTER,
        )

    embed = discord.Embed(
        title="📋 Панель подачі звітів",
        description=(
            "📋 **Інструкція з подання звіту за контракт:**\n\n"
            "**Шановні учасники! 👋**\n"
            "Просимо вас уважно та відповідально підходити до заповнення звітів, адже від правильності введених даних залежить їх схвалення та нарахування заробітної плати.\n\n"
            "**🔹 Покрокова подача звіту:**\n\n"
            "1️⃣ Натисніть кнопку «Подати звіт»\n"
            "2️⃣ У чаті з’явиться меню — оберіть пункт «Оберіть контракт»\n"
            "3️⃣ Виберіть контракт, який ви виконували\n"
            "4️⃣ Оберіть кількість учасників, які брали участь\n"
            "5️⃣ Натисніть кнопку «Другий етап»\n"
            "6️⃣ У вас відкриється меню для введення учасників\n\n"
            "**👥 Як правильно вказувати учасників:**\n\n"
            "У відповідному полі потрібно вказати **ВСІХ** гравців, які брали участь у контракті (включаючи себе).\n\n"
            "**📌 Формат (обов’язковий):**\n"
            "`Нікнейм Гравця #ID`\n\n"
            "**📌 Приклади:**\n\n"
            "• **1 учасник:**\n"
            "`Romeo Hunter #19857`\n\n"
            "• **2 учасники:**\n"
            "`Romeo Hunter #19857, Vlados Hunter #4589`\n\n"
            "• **3+ учасників** — аналогічно, через кому\n\n"
            "**❗ ВАЖЛИВІ ПРАВИЛА:**\n\n"
            "• Між нікнеймами **ЗАВЖДИ** має бути кома\n"
            "• Після останнього гравця кому ставити **НЕ** потрібно\n"
            "• Пишіть ніки без помилок\n"
            "• Вказуйте лише реальних учасників\n"
            "• Обов’язково зазначайте всіх учасників без винятку\n\n"
            "**📸 ВИМОГИ ДО СКРІНШОТІВ:**\n\n"
            "До звіту **ОБОВ’ЯЗКОВО** потрібно додати **2 посилання на скріншоти:**\n\n"
            "**🔹 Перше посилання — початок контракту:**\n\n"
            "Скріншот, де видно що ви разом стоїте (строєм)\n\n"
            "І на екрані є повідомлення про активацію контракту\n\n"
            "**📌 Якщо не встигли зробити цей момент:**\n"
            "можна завантажити в одне посилання:\n\n"
            "скрін, де ви стоїте разом\n\n"
            "скрін з планшету, де видно що контракт активовано\n\n"
            "**🔹 Друге посилання — завершення контракту:**\n\n"
            "Скріншот, де видно що ви разом\n\n"
            "І в чаті написано, що ваша організація виконала контракт\n\n"
            "**📌 Якщо не встигли:**\n"
            "можна додати:\n\n"
            "скрін з планшету контрактів\n\n"
            "або інше підтвердження виконання\n\n"
            "**⚠️ Зверніть увагу:**\n\n"
            "У разі:\n"
            "• неправдивої інформації\n"
            "• помилок у ніках або ID\n"
            "• відсутності когось з учасників\n"
            "• відсутності або неправильних скріншотів\n\n"
            "➡️ звіт **НЕ** буде схвалений\n"
            "➡️ заробітна плата **НЕ** нараховується\n\n"
            "Дякуємо за вашу уважність та відповідальність! 🤝\n\n"
            "*HUNTER | Contract System*"
        ),
        color=0x2ECC71,
    )

    await interaction.channel.send(embed=embed, view=MainView())
    await interaction.response.send_message(
        "✅ Панель створена",
        ephemeral=True,
        delete_after=DELETE_AFTER,
    )


# ========================= ERROR LOGGING =========================
@bot.tree.error
async def on_app_command_error(interaction: discord.Interaction, error: discord.app_commands.AppCommandError):
    print("APP COMMAND ERROR:", repr(error))
    traceback.print_exc()

    if interaction.response.is_done():
        await interaction.followup.send(
            "❌ Сталася помилка. Перевір консоль бота.",
            ephemeral=True,
            delete_after=DELETE_AFTER,
        )
    else:
        await interaction.response.send_message(
            "❌ Сталася помилка. Перевір консоль бота.",
            ephemeral=True,
            delete_after=DELETE_AFTER,
        )


# ========================= READY =========================
@bot.event
async def on_ready():
    await bot.tree.sync()
    print(f"BOT ONLINE: {bot.user}")


if not TOKEN:
    raise ValueError("Set DISCORD_BOT_TOKEN before starting the bot.")

bot.run(TOKEN)
